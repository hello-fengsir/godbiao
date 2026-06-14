"""导入产品参数到数据库 — 对齐当前表结构"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

from models import init_all_tables, get_db, bid_parameters_all

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "product_data.xlsx")


def import_products():
    init_all_tables()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    with get_db() as db:
        count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            product_line = sheet_name.strip()

            # 确保产品线存在（归属到第一个类别，默认"网络安全"）
            cat = db.execute("SELECT id FROM bid_categories LIMIT 1").fetchone()
            if not cat:
                cat_id = db.execute(
                    "INSERT INTO bid_categories (name, sort_order) VALUES ('网络安全', 1)"
                ).lastrowid
            else:
                cat_id = cat[0]

            line = db.execute(
                "SELECT id FROM bid_product_lines WHERE name=? AND category_id=?",
                (product_line, cat_id)
            ).fetchone()
            if not line:
                line_id = db.execute(
                    "INSERT INTO bid_product_lines (category_id, name) VALUES (?,?)",
                    (cat_id, product_line)
                ).lastrowid
            else:
                line_id = line[0]

            max_so = db.execute(
                "SELECT MAX(sort_order) FROM bid_parameters WHERE line_id=?", (line_id,)
            ).fetchone()[0] or 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                vals = [str(c).strip() if c else "" for c in row]
                if not any(vals):
                    continue

                # 第一列=参数名称，第二列=参数值（可选）
                param_name = vals[0] if vals else ""
                param_value = vals[1] if len(vals) > 1 else ""
                if not param_name:
                    continue

                max_so += 1
                db.execute(
                    "INSERT INTO bid_parameters (line_id, param_type, param_name, param_value, sort_order) VALUES (?,?,?,?,?)",
                    (line_id, "software", param_name, param_value, max_so)
                )
                count += 1

        print(f"导入完成：{count} 条参数")


if __name__ == "__main__":
    import_products()
